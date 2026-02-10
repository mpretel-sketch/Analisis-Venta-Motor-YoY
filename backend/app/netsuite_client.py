"""
Cliente de NetSuite para conexión via RESTlet con autenticación TBA (OAuth 1.0a)
"""
from __future__ import annotations

import os
import json
import logging
from typing import Dict, List, Optional
from urllib.parse import urljoin

import requests
from requests_oauthlib import OAuth1
import pandas as pd

logger = logging.getLogger("netsuite_client")


class NetSuiteError(Exception):
    """Error de conexión o consulta a NetSuite"""
    pass


class NetSuiteClient:
    """
    Cliente para conectar con NetSuite via RESTlet usando Token-Based Authentication (OAuth 1.0a)
    """

    def __init__(
        self,
        account: str,
        consumer_key: str,
        consumer_secret: str,
        token_id: str,
        token_secret: str,
        restlet_url: Optional[str] = None,
    ):
        """
        Inicializa el cliente de NetSuite

        Args:
            account: Account ID de NetSuite (ej: "8085023")
            consumer_key: Consumer Key de la integración
            consumer_secret: Consumer Secret de la integración
            token_id: Token ID del usuario
            token_secret: Token Secret del usuario
            restlet_url: URL completa del RESTlet (opcional, se puede construir)
        """
        self.account = account.replace("_", "-").upper()
        self.consumer_key = consumer_key
        self.consumer_secret = consumer_secret
        self.token_id = token_id
        self.token_secret = token_secret
        self.restlet_url = restlet_url

        # Construir URL base si no se proporciona RESTlet URL
        if not self.restlet_url:
            # Nota: El usuario debe proporcionar la URL completa del RESTlet
            logger.warning("No se proporcionó URL del RESTlet. Debes configurar NS_RESTLET_URL en .env")

    def _get_oauth(self) -> OAuth1:
        """
        Crea el objeto OAuth1 para autenticación TBA

        Returns:
            OAuth1: Objeto de autenticación configurado
        """
        return OAuth1(
            client_key=self.consumer_key,
            client_secret=self.consumer_secret,
            resource_owner_key=self.token_id,
            resource_owner_secret=self.token_secret,
            realm=self.account,
            signature_method="HMAC-SHA256",
        )

    def fetch_sales_data(
        self,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        filters: Optional[Dict] = None,
    ) -> pd.DataFrame:
        """
        Obtiene datos de ventas desde NetSuite via RESTlet

        Args:
            start_date: Fecha inicio (formato YYYY-MM-DD) - opcional
            end_date: Fecha fin (formato YYYY-MM-DD) - opcional
            filters: Filtros adicionales para el RESTlet

        Returns:
            pd.DataFrame: DataFrame con columnas Cliente, Hotel - Code, Ubicación, y meses

        Raises:
            NetSuiteError: Si hay error en la consulta
        """
        if not self.restlet_url:
            raise NetSuiteError(
                "No se ha configurado NS_RESTLET_URL. "
                "Debes proporcionar la URL completa del RESTlet en el archivo .env"
            )

        # Preparar parámetros para el RESTlet
        params = {}
        if start_date:
            params["startDate"] = start_date
        if end_date:
            params["endDate"] = end_date
        if filters:
            params.update(filters)

        logger.info(f"Consultando RESTlet de NetSuite: {self.restlet_url}")
        logger.debug(f"Parámetros: {params}")

        try:
            # Hacer petición GET al RESTlet con autenticación OAuth
            response = requests.get(
                self.restlet_url,
                auth=self._get_oauth(),
                headers={
                    "Content-Type": "application/json",
                    "Accept": "application/json",
                },
                params=params,
                timeout=120,  # 2 minutos timeout para queries grandes
            )

            response.raise_for_status()

            # Parsear respuesta JSON
            data = response.json()

            # Validar que tengamos datos
            if not data or not isinstance(data, list):
                raise NetSuiteError("El RESTlet no devolvió datos válidos (esperado: array de objetos)")

            if len(data) == 0:
                logger.warning("El RESTlet devolvió 0 registros")
                return pd.DataFrame()

            # Convertir a DataFrame
            df = pd.DataFrame(data)

            # Validar columnas requeridas
            required_cols = ["Cliente"]
            missing = [col for col in required_cols if col not in df.columns]
            if missing:
                raise NetSuiteError(
                    f"El RESTlet no devolvió las columnas requeridas. Faltantes: {missing}. "
                    f"Columnas recibidas: {list(df.columns)}"
                )

            logger.info(f"✓ Datos obtenidos: {len(df)} registros, {len(df.columns)} columnas")

            return df

        except requests.exceptions.HTTPError as exc:
            error_msg = f"Error HTTP al consultar NetSuite: {exc}"
            try:
                error_detail = exc.response.json()
                error_msg += f"\nDetalle: {error_detail}"
            except Exception:
                error_msg += f"\nRespuesta: {exc.response.text[:500]}"
            logger.error(error_msg)
            raise NetSuiteError(error_msg) from exc

        except requests.exceptions.Timeout as exc:
            error_msg = "Timeout al consultar NetSuite. El RESTlet tardó más de 2 minutos."
            logger.error(error_msg)
            raise NetSuiteError(error_msg) from exc

        except requests.exceptions.RequestException as exc:
            error_msg = f"Error de conexión con NetSuite: {exc}"
            logger.error(error_msg)
            raise NetSuiteError(error_msg) from exc

        except Exception as exc:
            error_msg = f"Error inesperado al procesar datos de NetSuite: {exc}"
            logger.error(error_msg)
            raise NetSuiteError(error_msg) from exc

    def test_connection(self) -> Dict[str, any]:
        """
        Prueba la conexión con NetSuite

        Returns:
            Dict con resultado de la prueba
        """
        try:
            # Intentar fetch con límite pequeño para prueba rápida
            df = self.fetch_sales_data()
            return {
                "success": True,
                "message": f"Conexión exitosa. {len(df)} registros disponibles.",
                "records": len(df),
            }
        except Exception as exc:
            return {
                "success": False,
                "message": str(exc),
            }


def get_netsuite_client() -> NetSuiteClient:
    """
    Crea una instancia del cliente de NetSuite usando variables de entorno

    Returns:
        NetSuiteClient: Cliente configurado

    Raises:
        ValueError: Si faltan variables de entorno requeridas
    """
    account = os.getenv("NS_ACCOUNT")
    consumer_key = os.getenv("NS_CONSUMER_KEY")
    consumer_secret = os.getenv("NS_CONSUMER_SECRET")
    token_id = os.getenv("NS_TOKEN_ID")
    token_secret = os.getenv("NS_TOKEN_SECRET")
    restlet_url = os.getenv("NS_RESTLET_URL")

    missing = []
    if not account:
        missing.append("NS_ACCOUNT")
    if not consumer_key:
        missing.append("NS_CONSUMER_KEY")
    if not consumer_secret:
        missing.append("NS_CONSUMER_SECRET")
    if not token_id:
        missing.append("NS_TOKEN_ID")
    if not token_secret:
        missing.append("NS_TOKEN_SECRET")
    if not restlet_url:
        missing.append("NS_RESTLET_URL")

    if missing:
        raise ValueError(
            f"Faltan variables de entorno requeridas para NetSuite: {', '.join(missing)}. "
            "Revisa tu archivo .env"
        )

    return NetSuiteClient(
        account=account,
        consumer_key=consumer_key,
        consumer_secret=consumer_secret,
        token_id=token_id,
        token_secret=token_secret,
        restlet_url=restlet_url,
    )


def dataframe_to_excel_format(df: pd.DataFrame) -> bytes:
    """
    Convierte el DataFrame de NetSuite al formato Excel esperado por analyze_yoy

    Args:
        df: DataFrame con datos de NetSuite

    Returns:
        bytes: Excel en formato bytes para pasarlo a analyze_yoy
    """
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos NetSuite"

    # Escribir encabezados
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Escribir datos
    for row_idx, row in enumerate(df.values, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Guardar en bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
