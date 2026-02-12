from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import os
import json
import hashlib
from collections import OrderedDict
from typing import Dict, List, Optional, Tuple
import re

import numpy as np
import pandas as pd
import requests


MONTH_MAP = {
    "ene": 1,
    "feb": 2,
    "mar": 3,
    "abr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dic": 12,
}

MONTH_REGEXES = [
    # ene 2024, ene-2024, ene_2024
    re.compile(r"^(?P<mon>ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)\s*[-_/]?\s*(?P<year>20\d{2})$", re.IGNORECASE),
    # 2024 ene, 2024-ene
    re.compile(r"^(?P<year>20\d{2})\s*[-_/]?\s*(?P<mon>ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)$", re.IGNORECASE),
    # ene 24
    re.compile(r"^(?P<mon>ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)\s*[-_/]?\s*(?P<year>\d{2})$", re.IGNORECASE),
]

MONTH_NAME = {
    1: "ene",
    2: "feb",
    3: "mar",
    4: "abr",
    5: "may",
    6: "jun",
    7: "jul",
    8: "ago",
    9: "sep",
    10: "oct",
    11: "nov",
    12: "dic",
}


@dataclass
class MonthColumn:
    col: str
    month: int
    year: int
    date: datetime


@dataclass
class YoYPair:
    current: MonthColumn
    previous: MonthColumn
    label: str


class AnalysisError(Exception):
    pass


_EXCEL_CACHE = OrderedDict()
_EXCEL_CACHE_MAX = 4


def _excel_cache_get(file_bytes: bytes) -> Optional[pd.DataFrame]:
    key = hashlib.md5(file_bytes).hexdigest()
    cached = _EXCEL_CACHE.get(key)
    if cached is None:
        return None
    _EXCEL_CACHE.move_to_end(key)
    return cached.copy(deep=True)


def _excel_cache_set(file_bytes: bytes, df: pd.DataFrame) -> None:
    key = hashlib.md5(file_bytes).hexdigest()
    _EXCEL_CACHE[key] = df
    _EXCEL_CACHE.move_to_end(key)
    while len(_EXCEL_CACHE) > _EXCEL_CACHE_MAX:
        _EXCEL_CACHE.popitem(last=False)


def _detect_header_row(df_head: pd.DataFrame) -> Optional[int]:
    for idx in range(min(15, len(df_head))):
        row = df_head.iloc[idx]
        if row.astype(str).str.strip().str.lower().eq("cliente").any():
            return idx
    return None


def _read_excel(file_bytes: bytes, filename: str) -> pd.DataFrame:
    cached_df = _excel_cache_get(file_bytes)
    if cached_df is not None:
        return cached_df

    ext = filename.lower().split(".")[-1]
    engine = None
    if ext == "xls":
        engine = "xlrd"
    elif ext == "xlsx":
        engine = "openpyxl"

    if engine is None:
        raise AnalysisError("Formato no soportado. Usa .xls o .xlsx")

    try:
        # Only sample first rows to detect header and avoid a full initial parse.
        preview = pd.read_excel(BytesIO(file_bytes), engine=engine, header=None, nrows=30)
    except Exception:
        # If extension lies (e.g., .xls but actually xlsx), fallback to openpyxl.
        if engine != "openpyxl":
            preview = pd.read_excel(BytesIO(file_bytes), engine="openpyxl", header=None, nrows=30)
            engine = "openpyxl"
        else:
            raise
    header_row = _detect_header_row(preview)
    if header_row is None:
        # Fallback to old format that starts at row 7 (index 6)
        header_row = 6

    df = pd.read_excel(BytesIO(file_bytes), engine=engine, header=header_row)
    _excel_cache_set(file_bytes, df)
    return df.copy(deep=True)


def _parse_month_column(col: str) -> Optional[MonthColumn]:
    col_str = str(col).strip().lower()
    if not col_str:
        return None

    for rx in MONTH_REGEXES:
        m = rx.match(col_str)
        if not m:
            continue
        mon = m.group("mon").lower()
        year_raw = m.group("year")
        year = int(year_raw)
        if year < 100:
            year += 2000
        month_num = MONTH_MAP.get(mon)
        if not month_num:
            return None
        return MonthColumn(col=str(col), month=month_num, year=year, date=datetime(year, month_num, 1))

    # Try split format: "ene 2024" with extra tokens
    parts = re.split(r"\s+", col_str)
    if len(parts) >= 2:
        mon = parts[0]
        year_part = parts[1]
        if mon in MONTH_MAP and re.match(r"^20\d{2}$", year_part):
            year = int(year_part)
            month_num = MONTH_MAP[mon]
            return MonthColumn(col=str(col), month=month_num, year=year, date=datetime(year, month_num, 1))

    return None


def _find_month_columns(df: pd.DataFrame) -> List[MonthColumn]:
    cols: List[MonthColumn] = []
    for col in df.columns:
        mc = _parse_month_column(col)
        if mc:
            cols.append(mc)
    return sorted(cols, key=lambda x: x.date)


def _find_latest_pair(month_cols: List[MonthColumn]) -> YoYPair:
    pairs: List[YoYPair] = []
    for mc in month_cols:
        prev = next((m for m in month_cols if m.month == mc.month and m.year == mc.year - 1), None)
        if prev:
            pairs.append(YoYPair(current=mc, previous=prev, label=f"{mc.col} vs {prev.col}"))

    if not pairs:
        raise AnalysisError("No se encontraron pares YoY. Se necesitan meses del mismo mes en años distintos.")

    return pairs[-1]

def _month_key(mc: MonthColumn) -> str:
    return f"{mc.year:04d}-{mc.month:02d}"


def _label_for_month(year: int, month: int) -> str:
    return f"{MONTH_NAME.get(month, str(month))} {year}"


def _find_month_by_key(month_cols: List[MonthColumn], key: str) -> Optional[MonthColumn]:
    for mc in month_cols:
        if _month_key(mc) == key:
            return mc
    return None


def _build_available_months(month_cols: List[MonthColumn]) -> List[Dict]:
    keys = {_month_key(mc): mc for mc in month_cols}
    available = []
    for mc in month_cols:
        prev_key = f"{mc.year - 1:04d}-{mc.month:02d}"
        has_prev = prev_key in keys
        available.append({
            "key": _month_key(mc),
            "label": _label_for_month(mc.year, mc.month),
            "year": mc.year,
            "month": mc.month,
            "hasPrev": has_prev,
        })
    return available


def _period_columns(
    month_cols: List[MonthColumn],
    mode: str,
    month_key: Optional[str],
) -> Tuple[List[MonthColumn], List[MonthColumn], str, str, str, str]:
    if not month_cols:
        raise AnalysisError("No se detectaron columnas de meses.")

    month_cols = sorted(month_cols, key=lambda x: x.date)
    if month_key:
        selected = _find_month_by_key(month_cols, month_key)
    else:
        selected = month_cols[-1]

    if not selected:
        raise AnalysisError("Mes seleccionado no disponible.")

    if mode == "month":
        prev = _find_month_by_key(month_cols, f"{selected.year - 1:04d}-{selected.month:02d}")
        if not prev:
            raise AnalysisError("No existe el mismo mes del año anterior para el mes seleccionado.")
        label = f"{selected.col} vs {prev.col}"
        period_label = f"{_label_for_month(selected.year, selected.month)} vs {_label_for_month(prev.year, prev.month)}"
        return [selected], [prev], label, period_label, selected.col, prev.col

    if mode == "ytd":
        curr_cols = [m for m in month_cols if m.year == selected.year and m.month <= selected.month]
        prev_cols = [m for m in month_cols if m.year == selected.year - 1 and m.month <= selected.month]
        if len(prev_cols) != len(curr_cols) or not prev_cols:
            raise AnalysisError("No hay suficientes meses del año anterior para YTD.")
        current_label = f"YTD {selected.year} (ene-{MONTH_NAME[selected.month]})"
        previous_label = f"YTD {selected.year - 1} (ene-{MONTH_NAME[selected.month]})"
        label = f"{current_label} vs {previous_label}"
        period_label = label
        return curr_cols, prev_cols, label, period_label, current_label, previous_label

    if mode in {"rolling3", "rolling6"}:
        window = 3 if mode == "rolling3" else 6
        idx = month_cols.index(selected)
        if idx + 1 < window:
            raise AnalysisError("No hay suficientes meses para el rolling seleccionado.")
        curr_window = month_cols[idx - window + 1 : idx + 1]
        prev_window = []
        for m in curr_window:
            prev = _find_month_by_key(month_cols, f"{m.year - 1:04d}-{m.month:02d}")
            if not prev:
                raise AnalysisError("No hay suficientes meses del año anterior para el rolling seleccionado.")
            prev_window.append(prev)
        current_label = f"Rolling {window}M hasta {_label_for_month(selected.year, selected.month)}"
        previous_label = f"Rolling {window}M hasta {_label_for_month(selected.year - 1, selected.month)}"
        label = f"{current_label} vs {previous_label}"
        period_label = label
        return curr_window, prev_window, label, period_label, current_label, previous_label

    raise AnalysisError("Modo no soportado. Usa month, ytd, rolling3 o rolling6.")


def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    if "Cliente" not in df.columns:
        raise AnalysisError("No se encontró la columna 'Cliente'. Revisa el archivo.")

    df = df[df["Cliente"].notna()].copy()
    df = df[~df["Cliente"].astype(str).str.strip().isin(["Ventas", "Total"])].copy()
    return df

def _build_hotel_series(df: pd.DataFrame, month_cols: List[MonthColumn], top_names: List[str]) -> Dict[str, List[Dict]]:
    series = {}
    if not top_names:
        return series

    for name in top_names:
        hotel_df = df[df["Cliente"] == name]
        if hotel_df.empty:
            continue
        row = hotel_df.iloc[0]
        items = []
        for mc in month_cols:
            prev = _find_month_by_key(month_cols, f"{mc.year - 1:04d}-{mc.month:02d}")
            if not prev:
                continue
            curr_val = float(pd.to_numeric(row.get(mc.col, 0), errors="coerce") or 0)
            prev_val = float(pd.to_numeric(row.get(prev.col, 0), errors="coerce") or 0)
            var_pct = ((curr_val - prev_val) / prev_val * 100) if prev_val > 0 else None
            items.append({
                "label": _label_for_month(mc.year, mc.month),
                "curr": curr_val,
                "varPct": var_pct,
            })
        series[name] = items
    return series


def analyze_yoy(
    file_bytes: bytes,
    filename: str,
    alert_threshold: float = -30.0,
    mode: str = "month",
    month_key: Optional[str] = None,
    search: Optional[str] = None,
    location: Optional[str] = None,
    impact_min: Optional[float] = None,
    impact_max: Optional[float] = None,
    var_min: Optional[float] = None,
    var_max: Optional[float] = None,
    persist_threshold: Optional[float] = None,
    recovery_threshold: Optional[float] = None,
    churn_months: int = 9,
) -> Dict:
    df = _read_excel(file_bytes, filename)
    df = _sanitize_df(df)

    month_cols = _find_month_columns(df)
    available_months = _build_available_months(month_cols)
    curr_cols, prev_cols, label, period_label, current_label, previous_label = _period_columns(month_cols, mode, month_key)
    latest_col = curr_cols[-1].col
    prev_col = prev_cols[-1].col

    for mc in curr_cols + prev_cols:
        df[mc.col] = pd.to_numeric(df[mc.col], errors="coerce").fillna(0)

    df["Current_Sum"] = df[[m.col for m in curr_cols]].sum(axis=1)
    df["Previous_Sum"] = df[[m.col for m in prev_cols]].sum(axis=1)

    df["Var_Absoluta"] = df["Current_Sum"] - df["Previous_Sum"]
    prev_nonzero = df["Previous_Sum"] != 0
    both_zero = (df["Previous_Sum"] == 0) & (df["Current_Sum"] == 0)
    df["Var_%"] = np.where(
        prev_nonzero,
        (df["Current_Sum"] - df["Previous_Sum"]) / df["Previous_Sum"] * 100,
        np.where(both_zero, 0.0, np.nan),
    )
    df["Impacto"] = df["Var_Absoluta"].abs()

    # Cluster derivado por prefijo de Cliente hasta ':'
    df["Cluster"] = df["Cliente"].astype(str).str.split(":").str[0].str.strip()

    # Country/Area heuristics (si existen)
    country_col = None
    for col in ["País", "Pais", "Country", "Hotel Country", "Hotel Country "]:
        if col in df.columns:
            country_col = col
            break

    def _apply_filters(dfx: pd.DataFrame) -> pd.DataFrame:
        filtered = dfx.copy()
        if search:
            needle = str(search).strip().lower()
            if needle:
                hay = (
                    filtered["Cliente"].astype(str).str.lower()
                    + " "
                    + filtered.get("Hotel - Code", "").astype(str).str.lower()
                )
                if "Ubicación" in filtered.columns:
                    hay = hay + " " + filtered["Ubicación"].astype(str).str.lower()
                if country_col and country_col in filtered.columns:
                    hay = hay + " " + filtered[country_col].astype(str).str.lower()
                filtered = filtered[hay.str.contains(needle, na=False)]
        if location and location != "all" and "Ubicación" in filtered.columns:
            filtered = filtered[filtered["Ubicación"] == location]
        if impact_min is not None:
            filtered = filtered[filtered["Impacto"] >= impact_min]
        if impact_max is not None:
            filtered = filtered[filtered["Impacto"] <= impact_max]
        if var_min is not None:
            filtered = filtered[(filtered["Var_%"].notna()) & (filtered["Var_%"] >= var_min)]
        if var_max is not None:
            filtered = filtered[(filtered["Var_%"].notna()) & (filtered["Var_%"] <= var_max)]
        return filtered

    df_filtered = _apply_filters(df)

    def _safe_float(value):
        if value is None:
            return None
        if isinstance(value, float) and (np.isnan(value) or np.isinf(value)):
            return None
        return float(value)

    total_current = float(df_filtered["Current_Sum"].sum())
    total_previous = float(df_filtered["Previous_Sum"].sum())
    total_var = total_current - total_previous
    total_var_pct = (total_var / total_previous * 100) if total_previous > 0 else 0.0

    alerts = df_filtered[df_filtered["Var_%"] < alert_threshold].sort_values("Impacto", ascending=False)
    growth = df_filtered[(df_filtered["Var_Absoluta"] > 3000) | (df_filtered["Var_%"] > 50)].sort_values("Var_Absoluta", ascending=False)
    new_hotels = df_filtered[(df_filtered["Previous_Sum"] == 0) & (df_filtered["Current_Sum"] > 0)].sort_values("Current_Sum", ascending=False)
    lost_hotels = df_filtered[(df_filtered["Previous_Sum"] > 0) & (df_filtered["Current_Sum"] == 0)].sort_values("Previous_Sum", ascending=False)

    def _rows(dfx: pd.DataFrame) -> List[Dict]:
        if dfx.empty:
            return []
        out = pd.DataFrame({
            "Cliente": dfx["Cliente"],
            "HotelCode": dfx["Hotel - Code"] if "Hotel - Code" in dfx.columns else None,
            "Ubicacion": dfx["Ubicación"] if "Ubicación" in dfx.columns else None,
            "Prev": dfx["Previous_Sum"],
            "Curr": dfx["Current_Sum"],
            "VarAbs": dfx["Var_Absoluta"],
            "VarPct": dfx["Var_%"],
        })
        out = out.replace([np.inf, -np.inf], np.nan)
        out["Prev"] = out["Prev"].fillna(0.0).astype(float)
        out["Curr"] = out["Curr"].fillna(0.0).astype(float)
        out["VarAbs"] = out["VarAbs"].fillna(0.0).astype(float)
        out["VarPct"] = out["VarPct"].astype(float)
        out = out.where(pd.notna(out), None)
        return out.to_dict("records")

    ubicacion_analysis = None
    if "Ubicación" in df_filtered.columns:
        ubicacion_analysis = df_filtered.groupby("Ubicación").agg({
            "Previous_Sum": "sum",
            "Current_Sum": "sum",
            "Var_Absoluta": "sum",
        })
        ubicacion_analysis["Var_%"] = np.where(
            ubicacion_analysis["Previous_Sum"] != 0,
            (ubicacion_analysis["Current_Sum"] / ubicacion_analysis["Previous_Sum"] - 1) * 100,
            np.nan,
        )
        ubicacion_analysis = ubicacion_analysis.sort_values("Var_Absoluta", ascending=False)

    locations = []
    if ubicacion_analysis is not None:
        for ubicacion, row in ubicacion_analysis.iterrows():
            locations.append({
                "Ubicacion": ubicacion,
                "Prev": _safe_float(row["Previous_Sum"]) or 0.0,
                "Curr": _safe_float(row["Current_Sum"]) or 0.0,
                "VarAbs": _safe_float(row["Var_Absoluta"]) or 0.0,
                "VarPct": _safe_float(row["Var_%"]),
            })

    # Serie temporal para gráficos (totales mensuales con YoY disponible)
    series = []
    for mc in month_cols:
        prev = _find_month_by_key(month_cols, f"{mc.year - 1:04d}-{mc.month:02d}")
        if not prev:
            continue
        curr_total = float(df_filtered[mc.col].sum())
        prev_total = float(df_filtered[prev.col].sum())
        var_pct = ((curr_total - prev_total) / prev_total * 100) if prev_total > 0 else None
        series.append({
            "key": _month_key(mc),
            "label": _label_for_month(mc.year, mc.month),
            "curr": curr_total,
            "prev": prev_total,
            "varPct": var_pct,
        })

    # Sparklines for top 10 in alerts and growth
    top_alerts = [row["Cliente"] for row in alerts.head(10).to_dict("records")]
    top_growth = [row["Cliente"] for row in growth.head(10).to_dict("records")]
    hotel_series = {
        "alerts": _build_hotel_series(df_filtered, month_cols, top_alerts),
        "growth": _build_hotel_series(df_filtered, month_cols, top_growth),
    }

    persist_threshold = alert_threshold if persist_threshold is None else persist_threshold
    recovery_threshold = 0.0 if recovery_threshold is None else recovery_threshold

    # Alertas inteligentes: persistente y recuperación (basado en últimos 2 meses con YoY disponible)
    intelligent_alerts = {"persistent": [], "recovery": []}
    if len(month_cols) >= 2:
        latest = month_cols[-1]
        prev_month = month_cols[-2]
        latest_prev = _find_month_by_key(month_cols, f"{latest.year - 1:04d}-{latest.month:02d}")
        prev_prev = _find_month_by_key(month_cols, f"{prev_month.year - 1:04d}-{prev_month.month:02d}")
        if latest_prev and prev_prev:
            curr_last = pd.to_numeric(df_filtered[latest.col], errors="coerce").fillna(0.0)
            prev_last = pd.to_numeric(df_filtered[latest_prev.col], errors="coerce").fillna(0.0)
            curr_prev = pd.to_numeric(df_filtered[prev_month.col], errors="coerce").fillna(0.0)
            prev_prev_vals = pd.to_numeric(df_filtered[prev_prev.col], errors="coerce").fillna(0.0)

            var_last = np.where(prev_last != 0, (curr_last - prev_last) / prev_last * 100, np.nan)
            var_prev = np.where(prev_prev_vals != 0, (curr_prev - prev_prev_vals) / prev_prev_vals * 100, np.nan)

            alerts_df = pd.DataFrame({
                "Cliente": df_filtered["Cliente"],
                "Ubicacion": df_filtered["Ubicación"] if "Ubicación" in df_filtered.columns else None,
                "VarPctLast": var_last,
                "VarPctPrev": var_prev,
            }).replace([np.inf, -np.inf], np.nan)

            valid = alerts_df["VarPctLast"].notna() & alerts_df["VarPctPrev"].notna()
            persistent_mask = valid & (alerts_df["VarPctLast"] <= persist_threshold) & (alerts_df["VarPctPrev"] <= persist_threshold)
            recovery_mask = valid & (alerts_df["VarPctPrev"] <= persist_threshold) & (alerts_df["VarPctLast"] >= recovery_threshold)

            intelligent_alerts["persistent"] = alerts_df[persistent_mask].to_dict("records")
            intelligent_alerts["recovery"] = alerts_df[recovery_mask].to_dict("records")

    # Consolidación por cluster / país / área comercial
    cluster_summary = df_filtered.groupby("Cluster").agg({
        "Previous_Sum": "sum",
        "Current_Sum": "sum",
        "Var_Absoluta": "sum",
    })
    cluster_summary["Var_%"] = np.where(
        cluster_summary["Previous_Sum"] != 0,
        (cluster_summary["Current_Sum"] / cluster_summary["Previous_Sum"] - 1) * 100,
        np.nan,
    )
    cluster_rows = [
        {
            "Cluster": idx,
            "Prev": _safe_float(row["Previous_Sum"]) or 0.0,
            "Curr": _safe_float(row["Current_Sum"]) or 0.0,
            "VarAbs": _safe_float(row["Var_Absoluta"]) or 0.0,
            "VarPct": _safe_float(row["Var_%"]),
        }
        for idx, row in cluster_summary.sort_values("Var_Absoluta", ascending=False).iterrows()
    ]

    country_rows = []
    if country_col:
        country_summary = df_filtered.groupby(country_col).agg({
            "Previous_Sum": "sum",
            "Current_Sum": "sum",
            "Var_Absoluta": "sum",
        })
        country_summary["Var_%"] = np.where(
            country_summary["Previous_Sum"] != 0,
            (country_summary["Current_Sum"] / country_summary["Previous_Sum"] - 1) * 100,
            np.nan,
        )
        for idx, row in country_summary.sort_values("Var_Absoluta", ascending=False).iterrows():
            country_rows.append({
                "Country": idx,
                "Prev": _safe_float(row["Previous_Sum"]) or 0.0,
                "Curr": _safe_float(row["Current_Sum"]) or 0.0,
                "VarAbs": _safe_float(row["Var_Absoluta"]) or 0.0,
                "VarPct": _safe_float(row["Var_%"]),
            })
    # Churn: hoteles con 0 ventas por N meses
    churn_list = []
    if month_cols:
        month_col_names = [m.col for m in month_cols]
        sales = df_filtered[month_col_names].apply(pd.to_numeric, errors="coerce").fillna(0.0).to_numpy()
        active = sales > 0
        rev_active = active[:, ::-1]
        has_any = rev_active.any(axis=1)
        idx_from_latest = np.where(has_any, np.argmax(rev_active, axis=1), len(month_cols))
        months_inactive = idx_from_latest.astype(int)

        churn_df = pd.DataFrame({
            "Cliente": df_filtered["Cliente"],
            "Ubicacion": df_filtered["Ubicación"] if "Ubicación" in df_filtered.columns else None,
            "MonthsInactive": months_inactive,
        })
        churn_list = churn_df[churn_df["MonthsInactive"] >= churn_months].to_dict("records")

    # Cohortes: por primer mes con ventas
    cohort_map = {}
    if month_cols:
        for _, row in df_filtered.iterrows():
            first = None
            for mc in month_cols:
                val = pd.to_numeric(row.get(mc.col, 0), errors="coerce")
                if val and val > 0:
                    first = mc
                    break
            if not first:
                continue
            cohort_key = _month_key(first)
            cohort_map.setdefault(cohort_key, [])
            cohort_map[cohort_key].append(row)

    cohort_rows = []
    cohort_cols = [_month_key(m) for m in month_cols]
    for cohort_key, rows in cohort_map.items():
        size = len(rows)
        if size == 0:
            continue
        base_month = next(m for m in month_cols if _month_key(m) == cohort_key)
        base_rev = sum(float(pd.to_numeric(r.get(base_month.col, 0), errors="coerce") or 0) for r in rows)
        active = []
        revenue = []
        for mc in month_cols:
            if mc.date < base_month.date:
                active.append(None)
                revenue.append(None)
                continue
            active_count = 0
            rev = 0.0
            for r in rows:
                val = float(pd.to_numeric(r.get(mc.col, 0), errors="coerce") or 0)
                if val > 0:
                    active_count += 1
                rev += val
            active.append(round(active_count / size * 100, 1))
            revenue.append(round((rev / base_rev * 100) if base_rev > 0 else 0.0, 1))
        cohort_rows.append({
            "cohort": cohort_key,
            "size": size,
            "active": active,
            "revenue": revenue,
        })

    cohort_rows.sort(key=lambda r: r.get('cohort'))

    return {
        "meta": {
            "latestLabel": current_label,
            "previousLabel": previous_label,
            "pairLabel": label,
            "periodLabel": period_label,
            "alertThreshold": alert_threshold,
            "mode": mode,
            "monthKey": month_key or _month_key(curr_cols[-1]),
            "availableMonths": available_months,
            "filters": {
                "search": search or "",
                "location": location or "all",
                "impactMin": impact_min,
                "impactMax": impact_max,
                "varMin": var_min,
                "varMax": var_max,
            },
            "intelligentThresholds": {
                "persistent": persist_threshold,
                "recovery": recovery_threshold,
            },
            "churnMonths": churn_months,
        },
        "summary": {
            "totalPrev": total_previous,
            "totalCurr": total_current,
            "totalVar": total_var,
            "totalVarPct": total_var_pct,
            "alertsCount": int(len(alerts)),
            "alertsImpact": float(alerts["Var_Absoluta"].sum()) if len(alerts) else 0.0,
            "growthCount": int(len(growth)),
            "growthImpact": float(growth["Var_Absoluta"].sum()) if len(growth) else 0.0,
            "newCount": int(len(new_hotels)),
            "newRevenue": float(new_hotels["Current_Sum"].sum()) if len(new_hotels) else 0.0,
            "lostCount": int(len(lost_hotels)),
            "lostRevenue": float(lost_hotels["Previous_Sum"].sum()) if len(lost_hotels) else 0.0,
        },
        "tables": {
            "alerts": _rows(alerts),
            "growth": _rows(growth),
            "new": _rows(new_hotels),
            "lost": _rows(lost_hotels),
            "locations": locations,
        },
        "series": series,
        "hotelSeries": hotel_series,
        "intelligentAlerts": intelligent_alerts,
        "clusters": {
            "byCluster": cluster_rows,
            "byCountry": country_rows,
            "byArea": locations,
        },
        "churn": churn_list,
        "cohorts": {
            "columns": cohort_cols,
            "rows": cohort_rows,
        },
        "aiSummary": _build_ai_summary(
            {
                "totalPrev": total_previous,
                "totalCurr": total_current,
                "totalVar": total_var,
                "totalVarPct": total_var_pct,
                "alertsCount": int(len(alerts)),
                "growthCount": int(len(growth)),
            },
            alerts,
            growth,
            country_rows,
            locations,
            period_label,
        ),
    }




def _build_ai_summary(
    summary: Dict,
    alerts: pd.DataFrame,
    growth: pd.DataFrame,
    country_rows: List[Dict],
    location_rows: List[Dict],
    period_label: str,
) -> Dict:
    heuristic = _build_ai_summary_heuristic(
        summary, alerts, growth, country_rows, location_rows, period_label
    )
    llm, llm_error = _build_ai_summary_llm(
        summary, alerts, growth, country_rows, location_rows, period_label
    )
    if llm:
        llm["llmFallbackReason"] = None
        return llm
    heuristic["llmFallbackReason"] = llm_error
    return heuristic


def _build_ai_summary_heuristic(
    summary: Dict,
    alerts: pd.DataFrame,
    growth: pd.DataFrame,
    country_rows: List[Dict],
    location_rows: List[Dict],
    period_label: str,
) -> Dict:
    total_var_pct = float(summary.get("totalVarPct", 0.0) or 0.0)
    total_var = float(summary.get("totalVar", 0.0) or 0.0)
    headline = (
        f"{period_label}: crecimiento YoY de {total_var_pct:.1f}% ({total_var:,.0f} EUR)."
        if total_var_pct >= 0
        else f"{period_label}: caida YoY de {abs(total_var_pct):.1f}% ({abs(total_var):,.0f} EUR)."
    )

    conclusions = []
    observations = []
    actions = []

    conclusions.append(
        f"Facturacion actual {summary.get('totalCurr', 0):,.0f} EUR vs {summary.get('totalPrev', 0):,.0f} EUR del periodo comparable."
    )

    if len(alerts):
        top_alert = alerts.iloc[0]
        conclusions.append(
            f"Mayor caida: {top_alert.get('Cliente', 'N/D')} con impacto {float(top_alert.get('Var_Absoluta', 0) or 0):,.0f} EUR."
        )
    if len(growth):
        top_growth = growth.iloc[0]
        conclusions.append(
            f"Mayor crecimiento: {top_growth.get('Cliente', 'N/D')} con mejora {float(top_growth.get('Var_Absoluta', 0) or 0):,.0f} EUR."
        )

    if country_rows:
        top_country = country_rows[0]
        observations.append(
            f"Pais lider por facturacion: {top_country.get('Country', 'N/D')} ({float(top_country.get('Curr', 0) or 0):,.0f} EUR)."
        )
    if location_rows:
        top_location = location_rows[0]
        observations.append(
            f"Ubicacion lider: {top_location.get('Ubicacion', 'N/D')} ({float(top_location.get('Curr', 0) or 0):,.0f} EUR)."
        )

    if summary.get("alertsCount", 0) > 0:
        actions.append("Priorizar plan de recuperacion en hoteles de mayor impacto negativo y monitorizar semanalmente.")
    if summary.get("growthCount", 0) > 0:
        actions.append("Replicar palancas comerciales de los hoteles con mayor crecimiento en clusters comparables.")
    if not actions:
        actions.append("Mantener seguimiento mensual y revisar elasticidad por ubicacion y segmento.")

    risks = []
    opportunities = []
    if len(alerts):
        risks.append("Concentracion de caida en pocos hoteles con impacto alto en margen.")
    if summary.get("totalVarPct", 0) is not None and float(summary.get("totalVarPct", 0) or 0) < 0:
        risks.append("Tendencia YoY negativa en el agregado; revisar pricing y mix por mercado.")
    if len(growth):
        opportunities.append("Escalar practicas comerciales de hoteles lideres en crecimiento.")
    if country_rows:
        opportunities.append("Profundizar en el pais lider para capturar demanda incremental.")

    return {
        "source": "heuristic",
        "headline": headline,
        "conclusions": conclusions[:4],
        "observations": observations[:4],
        "risks": risks[:3],
        "opportunities": opportunities[:3],
        "actions": actions[:4],
    }


def _build_ai_summary_llm(
    summary: Dict,
    alerts: pd.DataFrame,
    growth: pd.DataFrame,
    country_rows: List[Dict],
    location_rows: List[Dict],
    period_label: str,
) -> Tuple[Optional[Dict], Optional[str]]:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None, "missing_openai_api_key"

    model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
    timeout = float(os.getenv("OPENAI_TIMEOUT_SECONDS", "12"))

    payload = {
        "period": period_label,
        "summary": {
            "totalPrev": float(summary.get("totalPrev", 0) or 0),
            "totalCurr": float(summary.get("totalCurr", 0) or 0),
            "totalVar": float(summary.get("totalVar", 0) or 0),
            "totalVarPct": float(summary.get("totalVarPct", 0) or 0),
            "alertsCount": int(summary.get("alertsCount", 0) or 0),
            "growthCount": int(summary.get("growthCount", 0) or 0),
        },
        "topAlerts": [
            {
                "hotel": str(row.get("Cliente", "N/D")),
                "impact": float(row.get("Var_Absoluta", 0) or 0),
                "varPct": float(row.get("Var_%", 0) or 0),
            }
            for _, row in alerts.head(5).iterrows()
        ],
        "topGrowth": [
            {
                "hotel": str(row.get("Cliente", "N/D")),
                "impact": float(row.get("Var_Absoluta", 0) or 0),
                "varPct": float(row.get("Var_%", 0) or 0),
            }
            for _, row in growth.head(5).iterrows()
        ],
        "topCountries": country_rows[:5],
        "topLocations": location_rows[:5],
    }

    system_prompt = (
        "Eres un analista financiero senior de revenue hotelero. "
        "Devuelve SOLO JSON valido con esta estructura exacta: "
        '{"headline": string, "conclusions": string[<=4], "observations": string[<=4], "risks": string[<=3], "opportunities": string[<=3], "actions": string[<=4]}. '
        "Escribe en espanol neutro, directo y accionable para CFO."
    )

    try:
        response = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": model,
                "response_format": {"type": "json_object"},
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {
                        "role": "user",
                        "content": json.dumps(payload, ensure_ascii=True),
                    },
                ],
                "temperature": 0.2,
            },
            timeout=timeout,
        )
        response.raise_for_status()
        data = response.json()
        raw = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        parsed = json.loads(raw)
        if not isinstance(parsed, dict):
            return None, "invalid_llm_json"
        headline = str(parsed.get("headline", "")).strip()
        conclusions = [str(x).strip() for x in (parsed.get("conclusions") or []) if str(x).strip()]
        observations = [str(x).strip() for x in (parsed.get("observations") or []) if str(x).strip()]
        risks = [str(x).strip() for x in (parsed.get("risks") or []) if str(x).strip()]
        opportunities = [str(x).strip() for x in (parsed.get("opportunities") or []) if str(x).strip()]
        actions = [str(x).strip() for x in (parsed.get("actions") or []) if str(x).strip()]
        if not headline:
            return None, "empty_headline"
        return {
            "source": "llm",
            "headline": headline,
            "conclusions": conclusions[:4],
            "observations": observations[:4],
            "risks": risks[:3],
            "opportunities": opportunities[:3],
            "actions": actions[:4],
        }, None
    except Exception as exc:
        return None, str(exc)

def _safe_sheet_title(title: str) -> str:
    cleaned = "".join(ch for ch in title if ch not in "[]:*?/\\")
    return cleaned[:31] if len(cleaned) > 31 else cleaned


def _write_summary_sheet(wb, result: Dict, title: str):
    from openpyxl.styles import Font, PatternFill

    meta = result["meta"]
    summary = result["summary"]

    ws = wb.create_sheet(_safe_sheet_title(title))
    ws["A1"] = f"EARLY WARNING - {meta.get('periodLabel') or meta['pairLabel']}"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells("A1:D1")

    ws["A3"] = "MÉTRICAS PRINCIPALES"
    ws["A3"].font = Font(size=12, bold=True)

    metrics = [
        ["Facturación año anterior:", summary["totalPrev"]],
        ["Facturación año actual:", summary["totalCurr"]],
        ["Variación absoluta:", summary["totalVar"]],
        ["Variación %:", summary["totalVarPct"]],
        ["", ""],
        [f"Alertas (caídas >{abs(meta['alertThreshold'])}%):", summary["alertsCount"]],
        ["Impacto alertas:", summary["alertsImpact"]],
        ["Hoteles nuevos:", summary["newCount"]],
        ["Hoteles perdidos:", summary["lostCount"]],
    ]

    for i, (label, value) in enumerate(metrics, 4):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        if isinstance(value, (int, float)) and value != "":
            if "variación %" in label.lower():
                ws[f"B{i}"].number_format = "0.0%"
                ws[f"B{i}"].value = value / 100
            else:
                ws[f"B{i}"].number_format = "#,##0.00€"




def _write_clusters_sheet(wb, result: Dict, title: str, rows: List[Dict], label_key: str):
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet(_safe_sheet_title(title))
    headers = [label_key, result["meta"]["previousLabel"], result["meta"]["latestLabel"], "Variación €", "Variación %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")

    for r_idx, row in enumerate(rows, 2):
        values = [
            row.get(label_key),
            row.get("Prev"),
            row.get("Curr"),
            row.get("VarAbs"),
            row.get("VarPct"),
        ]
        for c_idx, value in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx in [2, 3, 4]:
                cell.number_format = "#,##0.00€"
            elif c_idx == 5:
                cell.number_format = "0.0%"
                if value is not None:
                    cell.value = value / 100

    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 20


def _write_churn_sheet(wb, result: Dict, title: str):
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet(_safe_sheet_title(title))
    headers = ["Hotel", "Ubicación", "Meses sin ventas"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="9B2C2C", end_color="9B2C2C", fill_type="solid")

    for r_idx, row in enumerate(result.get("churn", []), 2):
        ws.cell(row=r_idx, column=1, value=row.get("Cliente"))
        ws.cell(row=r_idx, column=2, value=row.get("Ubicacion"))
        ws.cell(row=r_idx, column=3, value=row.get("MonthsInactive"))

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18


def _write_cohorts_sheet(wb, result: Dict, title: str, metric: str):
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet(_safe_sheet_title(title))
    columns = ["Cohorte", "Tamaño"] + (result.get("cohorts", {}).get("columns") or [])
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")

    rows = result.get("cohorts", {}).get("rows") or []
    for r_idx, row in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=row.get("cohort"))
        ws.cell(row=r_idx, column=2, value=row.get("size"))
        values = row.get(metric) or []
        for c_idx, val in enumerate(values, 3):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if val is not None:
                cell.number_format = "0.0%"
                cell.value = val / 100

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    for col in range(3, len(columns) + 1):
        ws.column_dimensions[chr(64 + col)].width = 10
def _write_table_sheet(wb, result: Dict, title: str, rows: List[Dict], header_color: str):
    from openpyxl.styles import Font, PatternFill

    meta = result["meta"]
    latest_label = meta["latestLabel"]
    previous_label = meta["previousLabel"]

    headers = [
        "Hotel",
        "Code",
        "Ubicación",
        previous_label,
        latest_label,
        "Variación €",
        "Variación %",
    ]

    ws = wb.create_sheet(_safe_sheet_title(title))
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")

    for r_idx, row in enumerate(rows, 2):
        values = [
            row.get("Cliente"),
            row.get("HotelCode"),
            row.get("Ubicacion"),
            row.get("Prev"),
            row.get("Curr"),
            row.get("VarAbs"),
            row.get("VarPct"),
        ]
        for c_idx, value in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx in [4, 5, 6]:
                cell.number_format = "#,##0.00€"
            elif c_idx == 7:
                cell.number_format = "0.0%"
                if value is not None:
                    cell.value = value / 100

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 15


def _write_intelligent_sheet(wb, result: Dict, title: str):
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet(_safe_sheet_title(title))
    headers = ["Tipo", "Hotel", "Ubicación", "Mes actual %", "Mes previo %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3F3D56", end_color="3F3D56", fill_type="solid")

    row_idx = 2
    for item in result.get("intelligentAlerts", {}).get("persistent", []):
        ws.cell(row=row_idx, column=1, value="Persistente")
        ws.cell(row=row_idx, column=2, value=item.get("Cliente"))
        ws.cell(row=row_idx, column=3, value=item.get("Ubicacion"))
        ws.cell(row=row_idx, column=4, value=item.get("VarPctLast"))
        ws.cell(row=row_idx, column=5, value=item.get("VarPctPrev"))
        row_idx += 1

    for item in result.get("intelligentAlerts", {}).get("recovery", []):
        ws.cell(row=row_idx, column=1, value="Recuperación")
        ws.cell(row=row_idx, column=2, value=item.get("Cliente"))
        ws.cell(row=row_idx, column=3, value=item.get("Ubicacion"))
        ws.cell(row=row_idx, column=4, value=item.get("VarPctLast"))
        ws.cell(row=row_idx, column=5, value=item.get("VarPctPrev"))
        row_idx += 1

    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 22


def _build_workbook(result: Dict):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    _write_summary_sheet(wb, result, "Resumen Ejecutivo")
    _write_table_sheet(wb, result, "Alertas", result["tables"]["alerts"], "C00000")
    _write_table_sheet(wb, result, "Crecimientos", result["tables"]["growth"], "006100")
    if result.get("intelligentAlerts"):
        _write_intelligent_sheet(wb, result, "Inteligentes")
    if result.get("clusters"):
        _write_clusters_sheet(wb, result, "Clusters", result["clusters"]["byCluster"], "Cluster")
        if result["clusters"].get("byCountry"):
            _write_clusters_sheet(wb, result, "Paises", result["clusters"]["byCountry"], "Country")
        _write_table_sheet(wb, result, "Area Comercial", result["clusters"]["byArea"], "6B7280")
    if result.get("churn") is not None:
        _write_churn_sheet(wb, result, "Churn")
    if result.get("cohorts"):
        _write_cohorts_sheet(wb, result, "Cohortes Activos", "active")
        _write_cohorts_sheet(wb, result, "Cohortes Revenue", "revenue")
    return wb


def build_excel_report(
    file_bytes: bytes,
    filename: str,
    output_path: str,
    mode: str = "month",
    month_key: Optional[str] = None,
    search: Optional[str] = None,
    location: Optional[str] = None,
    impact_min: Optional[float] = None,
    impact_max: Optional[float] = None,
    var_min: Optional[float] = None,
    var_max: Optional[float] = None,
) -> Tuple[str, Dict]:
    result = analyze_yoy(
        file_bytes,
        filename,
        mode=mode,
        month_key=month_key,
        search=search,
        location=location,
        impact_min=impact_min,
        impact_max=impact_max,
        var_min=var_min,
        var_max=var_max,
    )
    wb = _build_workbook(result)
    wb.save(output_path)
    return output_path, result


def build_excel_report_bytes(
    file_bytes: bytes,
    filename: str,
    mode: str = "month",
    month_key: Optional[str] = None,
    search: Optional[str] = None,
    location: Optional[str] = None,
    impact_min: Optional[float] = None,
    impact_max: Optional[float] = None,
    var_min: Optional[float] = None,
    var_max: Optional[float] = None,
) -> Tuple[bytes, Dict]:
    from io import BytesIO

    result = analyze_yoy(
        file_bytes,
        filename,
        mode=mode,
        month_key=month_key,
        search=search,
        location=location,
        impact_min=impact_min,
        impact_max=impact_max,
        var_min=var_min,
        var_max=var_max,
    )
    wb = _build_workbook(result)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read(), result


def build_excel_report_bytes_multi(
    file_bytes: bytes,
    filename: str,
    requests: List[Dict],
    search: Optional[str] = None,
    location: Optional[str] = None,
    impact_min: Optional[float] = None,
    impact_max: Optional[float] = None,
    var_min: Optional[float] = None,
    var_max: Optional[float] = None,
) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    def _mode_tag(meta: Dict) -> str:
        mode = meta.get("mode", "month")
        mode_code = {
            "month": "M",
            "ytd": "YTD",
            "rolling3": "R3",
            "rolling6": "R6",
        }.get(mode, mode[:3].upper())
        key = meta.get("monthKey", "period")
        return f"{mode_code}-{key}"

    for req in requests:
        result = analyze_yoy(
            file_bytes,
            filename,
            mode=req.get("mode", "month"),
            month_key=req.get("monthKey"),
            search=search,
            location=location,
            impact_min=impact_min,
            impact_max=impact_max,
            var_min=var_min,
            var_max=var_max,
        )
        tag = req.get("label") or _mode_tag(result["meta"])
        _write_summary_sheet(wb, result, f"Resumen {tag}")
        _write_table_sheet(wb, result, f"Alertas {tag}", result["tables"]["alerts"], "C00000")
        _write_table_sheet(wb, result, f"Crec {tag}", result["tables"]["growth"], "006100")
        if result.get("clusters"):
            _write_clusters_sheet(wb, result, f"Clusters {tag}", result["clusters"]["byCluster"], "Cluster")
        if result.get("churn") is not None:
            _write_churn_sheet(wb, result, f"Churn {tag}")
        if result.get("cohorts"):
            _write_cohorts_sheet(wb, result, f"Cohortes A {tag}", "active")
            _write_cohorts_sheet(wb, result, f"Cohortes R {tag}", "revenue")

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


